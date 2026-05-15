from collections import defaultdict
from datetime import datetime, time, timezone
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo
from xml.sax.saxutils import escape

from flask import Blueprint, current_app, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table as PdfTable,
    TableStyle,
)

from app.models import ActionLog, Area, Customer, Table, TableSession, utc_now
from app.permissions import admin_required

reports_bp = Blueprint("admin_reports", __name__, url_prefix="/admin/reports")

DISPLAY_TIMEZONE = ZoneInfo("Europe/Istanbul")

COLOR_NAVY = "071D49"
COLOR_NAVY_SOFT = "0F2D67"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_BLUE = "EEF3FF"
COLOR_LIGHT_GRAY = "F4F7FB"
COLOR_BORDER = "DCE3EE"
COLOR_TEXT = "172033"
COLOR_MUTED = "6B7280"
COLOR_GREEN = "1F9D55"
COLOR_GREEN_SOFT = "E9F8EF"
COLOR_RED = "E63946"
COLOR_ORANGE = "F97316"

PDF_CONTENT_WIDTH = 27.35 * cm

ROLE_LABELS = {
    "admin": "Yönetici",
    "door_staff": "Kapı Personeli",
    "bar_staff": "Bar Personeli",
    "system": "Sistem",
    "-": "-",
}

TURKISH_ASCII_TRANSLATION = str.maketrans(
    {
        "ç": "c",
        "Ç": "C",
        "ğ": "g",
        "Ğ": "G",
        "ı": "i",
        "I": "I",
        "İ": "I",
        "ö": "o",
        "Ö": "O",
        "ş": "s",
        "Ş": "S",
        "ü": "u",
        "Ü": "U",
    }
)


def pdf_hex(hex_value):
    cleaned_value = str(hex_value).strip()

    if not cleaned_value.startswith("#"):
        cleaned_value = f"#{cleaned_value}"

    return colors.HexColor(cleaned_value)


def normalize_datetime_as_utc(value):
    if value is None:
        return None

    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)

    return value.astimezone(timezone.utc)


def get_today_local_date():
    return utc_now().astimezone(DISPLAY_TIMEZONE).date()


def parse_selected_date(value):
    if value is None:
        return get_today_local_date()

    cleaned_value = str(value).strip()

    if cleaned_value == "":
        return get_today_local_date()

    try:
        return datetime.strptime(cleaned_value, "%Y-%m-%d").date()
    except ValueError:
        return get_today_local_date()


def build_day_range_utc(selected_date):
    local_start = datetime.combine(selected_date, time.min).replace(tzinfo=DISPLAY_TIMEZONE)
    local_end = datetime.combine(selected_date, time.max).replace(tzinfo=DISPLAY_TIMEZONE)

    utc_start = local_start.astimezone(timezone.utc)
    utc_end = local_end.astimezone(timezone.utc)

    return utc_start, utc_end


def format_date_for_display(value):
    return value.strftime("%d.%m.%Y")


def format_datetime_for_display(value):
    normalized_value = normalize_datetime_as_utc(value)

    if normalized_value is None:
        return "-"

    local_value = normalized_value.astimezone(DISPLAY_TIMEZONE)

    return local_value.strftime("%H:%M")


def format_duration_from_minutes(duration_minutes):
    if duration_minutes is None:
        return "-"

    try:
        parsed_duration = int(duration_minutes)
    except (TypeError, ValueError):
        return "-"

    if parsed_duration < 1:
        return "0 dk"

    hours = parsed_duration // 60
    minutes = parsed_duration % 60

    if hours == 0:
        return f"{minutes} dk"

    if minutes == 0:
        return f"{hours} sa"

    return f"{hours} sa {minutes} dk"


def calculate_active_duration_minutes(check_in_at):
    normalized_check_in_at = normalize_datetime_as_utc(check_in_at)

    if normalized_check_in_at is None:
        return 0

    duration_seconds = (utc_now() - normalized_check_in_at).total_seconds()
    duration_minutes = int(duration_seconds // 60)

    if duration_minutes < 0:
        return 0

    return duration_minutes


def get_role_label(role):
    return ROLE_LABELS.get(role, role or "-")


def safe_extra_data(action_log):
    if isinstance(action_log.extra_data, dict):
        return action_log.extra_data

    return {}


def get_sessions_started_in_day(day_start_utc, day_end_utc):
    return (
        TableSession.query
        .join(Table, TableSession.table_id == Table.id)
        .join(Area, Table.area_id == Area.id)
        .filter(TableSession.check_in_at >= day_start_utc)
        .filter(TableSession.check_in_at <= day_end_utc)
        .order_by(TableSession.check_in_at.asc(), TableSession.id.asc())
        .all()
    )


def get_sessions_completed_in_day(day_start_utc, day_end_utc):
    return (
        TableSession.query
        .join(Table, TableSession.table_id == Table.id)
        .join(Area, Table.area_id == Area.id)
        .filter(TableSession.status == TableSession.STATUS_COMPLETED)
        .filter(TableSession.check_out_at.isnot(None))
        .filter(TableSession.check_out_at >= day_start_utc)
        .filter(TableSession.check_out_at <= day_end_utc)
        .order_by(TableSession.check_out_at.asc(), TableSession.id.asc())
        .all()
    )


def get_active_sessions():
    return (
        TableSession.query
        .join(Table, TableSession.table_id == Table.id)
        .join(Area, Table.area_id == Area.id)
        .filter(TableSession.status == TableSession.STATUS_ACTIVE)
        .order_by(TableSession.check_in_at.asc(), TableSession.id.asc())
        .all()
    )


def get_action_logs_in_day(day_start_utc, day_end_utc):
    return (
        ActionLog.query
        .filter(ActionLog.created_at >= day_start_utc)
        .filter(ActionLog.created_at <= day_end_utc)
        .order_by(ActionLog.created_at.asc(), ActionLog.id.asc())
        .all()
    )


def calculate_average_duration(completed_sessions):
    durations = [
        session.duration_minutes
        for session in completed_sessions
        if session.duration_minutes is not None
    ]

    if not durations:
        return None

    return round(sum(durations) / len(durations))


def build_area_summary_rows(started_sessions, completed_sessions):
    area_map = defaultdict(
        lambda: {
            "area_name": "",
            "session_count": 0,
            "guest_count": 0,
            "completed_count": 0,
            "total_duration_minutes": 0,
        }
    )

    for session in started_sessions:
        area_name = session.table.area.name
        area_map[area_name]["area_name"] = area_name
        area_map[area_name]["session_count"] += 1
        area_map[area_name]["guest_count"] += session.party_size or 0

    for session in completed_sessions:
        area_name = session.table.area.name
        area_map[area_name]["area_name"] = area_name
        area_map[area_name]["completed_count"] += 1
        area_map[area_name]["total_duration_minutes"] += session.duration_minutes or 0

    rows = []

    for area_data in area_map.values():
        completed_count = area_data["completed_count"]
        average_duration_minutes = None

        if completed_count > 0:
            average_duration_minutes = round(
                area_data["total_duration_minutes"] / completed_count
            )

        rows.append(
            {
                "area_name": area_data["area_name"],
                "session_count": area_data["session_count"],
                "guest_count": area_data["guest_count"],
                "completed_count": completed_count,
                "total_duration_minutes": area_data["total_duration_minutes"],
                "average_duration_minutes": average_duration_minutes,
                "total_duration": format_duration_from_minutes(
                    area_data["total_duration_minutes"]
                ),
                "average_duration": format_duration_from_minutes(
                    average_duration_minutes
                ),
            }
        )

    rows.sort(
        key=lambda row: (
            row["session_count"],
            row["guest_count"],
            row["area_name"],
        ),
        reverse=True,
    )

    return rows


def build_table_summary_rows(started_sessions, completed_sessions):
    table_map = defaultdict(
        lambda: {
            "table_code": "",
            "area_name": "",
            "session_count": 0,
            "guest_count": 0,
            "completed_count": 0,
            "total_duration_minutes": 0,
        }
    )

    for session in started_sessions:
        table_code = session.table.code
        table_map[table_code]["table_code"] = table_code
        table_map[table_code]["area_name"] = session.table.area.name
        table_map[table_code]["session_count"] += 1
        table_map[table_code]["guest_count"] += session.party_size or 0

    for session in completed_sessions:
        table_code = session.table.code
        table_map[table_code]["table_code"] = table_code
        table_map[table_code]["area_name"] = session.table.area.name
        table_map[table_code]["completed_count"] += 1
        table_map[table_code]["total_duration_minutes"] += session.duration_minutes or 0

    rows = []

    for table_data in table_map.values():
        completed_count = table_data["completed_count"]
        average_duration_minutes = None

        if completed_count > 0:
            average_duration_minutes = round(
                table_data["total_duration_minutes"] / completed_count
            )

        rows.append(
            {
                "table_code": table_data["table_code"],
                "area_name": table_data["area_name"],
                "session_count": table_data["session_count"],
                "guest_count": table_data["guest_count"],
                "completed_count": completed_count,
                "total_duration_minutes": table_data["total_duration_minutes"],
                "average_duration_minutes": average_duration_minutes,
                "total_duration": format_duration_from_minutes(
                    table_data["total_duration_minutes"]
                ),
                "average_duration": format_duration_from_minutes(
                    average_duration_minutes
                ),
            }
        )

    rows.sort(
        key=lambda row: (
            row["session_count"],
            row["guest_count"],
            row["table_code"],
        ),
        reverse=True,
    )

    return rows


def build_staff_summary_rows(action_logs):
    operation_types = {
        "table_assigned": "assigned_count",
        "table_cleared": "cleared_count",
        "table_transferred": "transfer_count",
    }

    staff_map = defaultdict(
        lambda: {
            "username": "Sistem",
            "role": "-",
            "assigned_count": 0,
            "cleared_count": 0,
            "transfer_count": 0,
            "total_count": 0,
        }
    )

    for action_log in action_logs:
        if action_log.action_type not in operation_types:
            continue

        username = action_log.username_snapshot or "Sistem"
        role = action_log.role_snapshot or "-"

        staff_map[username]["username"] = username
        staff_map[username]["role"] = get_role_label(role)
        staff_map[username][operation_types[action_log.action_type]] += 1
        staff_map[username]["total_count"] += 1

    rows = list(staff_map.values())
    rows.sort(
        key=lambda row: (
            row["total_count"],
            row["assigned_count"],
            row["cleared_count"],
            row["transfer_count"],
            row["username"],
        ),
        reverse=True,
    )

    return rows


def build_transfer_rows(action_logs):
    rows = []

    for action_log in action_logs:
        if action_log.action_type != "table_transferred":
            continue

        extra_data = safe_extra_data(action_log)

        rows.append(
            {
                "time": format_datetime_for_display(action_log.created_at),
                "old_table_code": extra_data.get("old_table_code", "-"),
                "old_area_name": extra_data.get("old_area_name", "-"),
                "new_table_code": extra_data.get("new_table_code", "-"),
                "new_area_name": extra_data.get("new_area_name", "-"),
                "party_size": extra_data.get("party_size", "-"),
                "customer_name": extra_data.get("customer_name") or "-",
                "username": action_log.username_snapshot or "Sistem",
            }
        )

    rows.sort(key=lambda row: row["time"])

    return rows


def build_session_rows(started_sessions):
    rows = []

    for session in started_sessions:
        if session.status == TableSession.STATUS_ACTIVE:
            status_label = "Aktif"
            duration_minutes = calculate_active_duration_minutes(session.check_in_at)
            duration_text = format_duration_from_minutes(duration_minutes)
            check_out_text = "-"
        elif session.status == TableSession.STATUS_COMPLETED:
            status_label = "Tamamlandı"
            duration_minutes = session.duration_minutes
            duration_text = format_duration_from_minutes(session.duration_minutes)
            check_out_text = format_datetime_for_display(session.check_out_at)
        else:
            status_label = "İptal"
            duration_minutes = None
            duration_text = "-"
            check_out_text = format_datetime_for_display(session.check_out_at)

        rows.append(
            {
                "table_code": session.table.code,
                "area_name": session.table.area.name,
                "party_size": session.party_size,
                "customer_name": session.customer_name or "-",
                "customer_phone": session.customer_phone or "-",
                "check_in": format_datetime_for_display(session.check_in_at),
                "check_out": check_out_text,
                "duration_minutes": duration_minutes,
                "duration": duration_text,
                "status_label": status_label,
            }
        )

    return rows


def find_most_used_area(area_rows):
    if not area_rows:
        return "-"

    first_row = area_rows[0]

    return (
        f"{first_row['area_name']} "
        f"({first_row['session_count']} oturum / {first_row['guest_count']} kişi)"
    )


def find_most_used_table(table_rows):
    if not table_rows:
        return "-"

    first_row = table_rows[0]

    return (
        f"{first_row['table_code']} "
        f"({first_row['session_count']} oturum / {first_row['guest_count']} kişi)"
    )


def build_customer_memory_summary(started_sessions):
    phone_session_count = 0
    linked_session_count = 0
    unique_customer_ids = set()
    unique_customers = {}

    for session in started_sessions:
        if session.customer_phone:
            phone_session_count += 1

        if session.customer_id:
            linked_session_count += 1
            unique_customer_ids.add(session.customer_id)

            if session.customer_id not in unique_customers:
                unique_customers[session.customer_id] = session.customer

    returning_customer_count = 0

    for customer in unique_customers.values():
        if customer is not None and (customer.visit_count or 0) > 1:
            returning_customer_count += 1

    total_registered_customer_count = Customer.query.count()

    return {
        "phone_session_count": phone_session_count,
        "linked_session_count": linked_session_count,
        "unique_customer_count": len(unique_customer_ids),
        "returning_customer_count": returning_customer_count,
        "total_registered_customer_count": total_registered_customer_count,
    }


def build_daily_summary_report(selected_date):
    day_start_utc, day_end_utc = build_day_range_utc(selected_date)

    started_sessions = get_sessions_started_in_day(day_start_utc, day_end_utc)
    completed_sessions = get_sessions_completed_in_day(day_start_utc, day_end_utc)
    active_sessions = get_active_sessions()
    action_logs = get_action_logs_in_day(day_start_utc, day_end_utc)

    transfer_logs = [
        action_log
        for action_log in action_logs
        if action_log.action_type == "table_transferred"
    ]

    total_guest_count = sum(
        session.party_size or 0
        for session in started_sessions
    )

    average_duration_minutes = calculate_average_duration(completed_sessions)

    area_rows = build_area_summary_rows(started_sessions, completed_sessions)
    table_rows = build_table_summary_rows(started_sessions, completed_sessions)
    staff_rows = build_staff_summary_rows(action_logs)
    transfer_rows = build_transfer_rows(action_logs)
    session_rows = build_session_rows(started_sessions)
    customer_memory_summary = build_customer_memory_summary(started_sessions)

    return {
        "selected_date": selected_date,
        "selected_date_value": selected_date.strftime("%Y-%m-%d"),
        "selected_date_display": format_date_for_display(selected_date),
        "generated_at": utc_now().astimezone(DISPLAY_TIMEZONE).strftime("%d.%m.%Y %H:%M"),
        "summary_cards": [
            {
                "label": "Bugün Açılan Oturum",
                "value": len(started_sessions),
                "hint": "Seçilen gün içinde masaya alınan oturum sayısı",
            },
            {
                "label": "Aktif Masa",
                "value": len(active_sessions),
                "hint": "Şu anda devam eden masa oturumları",
            },
            {
                "label": "Bugün Kapanan Oturum",
                "value": len(completed_sessions),
                "hint": "Seçilen gün içinde boşaltılan masa oturumu sayısı",
            },
            {
                "label": "Toplam Müşteri",
                "value": total_guest_count,
                "hint": "Seçilen gün açılan oturumlardaki kişi toplamı",
            },
            {
                "label": "Transfer",
                "value": len(transfer_logs),
                "hint": "Seçilen gün yapılan masa transferleri",
            },
            {
                "label": "Ortalama Süre",
                "value": format_duration_from_minutes(average_duration_minutes),
                "hint": "Bugün kapanan oturumlara göre ortalama süre",
            },
            {
                "label": "Telefonlu Oturum",
                "value": customer_memory_summary["phone_session_count"],
                "hint": "Telefon bilgisi girilen masa oturumları",
            },
            {
                "label": "Tekil Müşteri",
                "value": customer_memory_summary["unique_customer_count"],
                "hint": "Telefon numarasıyla müşteri hafızasına bağlanan kişi sayısı",
            },
            {
                "label": "Tekrar Gelen",
                "value": customer_memory_summary["returning_customer_count"],
                "hint": "Daha önce kaydı olan müşteriler",
            },
        ],
        "customer_memory_summary": customer_memory_summary,
        "most_used_area": find_most_used_area(area_rows),
        "most_used_table": find_most_used_table(table_rows),
        "area_rows": area_rows,
        "table_rows": table_rows[:10],
        "staff_rows": staff_rows,
        "transfer_rows": transfer_rows,
        "session_rows": session_rows,
    }


def style_title_area(worksheet, title, subtitle, total_columns):
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=total_columns,
    )
    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=total_columns,
    )

    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(size=18, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    subtitle_cell = worksheet.cell(row=2, column=1)
    subtitle_cell.value = subtitle
    subtitle_cell.font = Font(size=11, bold=False, color=COLOR_WHITE)
    subtitle_cell.fill = PatternFill("solid", fgColor=COLOR_NAVY_SOFT)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 22


def apply_common_sheet_style(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_margins.left = 0.35
    worksheet.page_margins.right = 0.35
    worksheet.page_margins.top = 0.45
    worksheet.page_margins.bottom = 0.45


def style_range_border(worksheet, min_row, max_row, min_col, max_col):
    thin_border = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_table(worksheet, start_row, headers, rows, column_widths=None):
    header_fill = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
    header_font = Font(bold=True, color=COLOR_NAVY)
    body_font = Font(color=COLOR_TEXT)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=column_index)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = start_row + 1

    if not rows:
        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(headers),
        )
        empty_cell = worksheet.cell(row=current_row, column=1)
        empty_cell.value = "Bu bölüm için veri bulunmuyor."
        empty_cell.font = Font(color=COLOR_MUTED, italic=True)
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        style_range_border(worksheet, start_row, current_row, 1, len(headers))
        return current_row + 2

    for row_values in rows:
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=current_row, column=column_index)
            cell.value = value
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        current_row += 1

    last_row = current_row - 1
    style_range_border(worksheet, start_row, last_row, 1, len(headers))
    worksheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"

    if column_widths:
        for column_index, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
    else:
        for column_index in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = 18

    return current_row + 2


def build_summary_sheet(workbook, report):
    worksheet = workbook.active
    worksheet.title = "Ozet"

    style_title_area(
        worksheet=worksheet,
        title="Lido Genel Gün Özeti",
        subtitle=f"Tarih: {report['selected_date_display']} | Oluşturma: {report['generated_at']}",
        total_columns=6,
    )

    apply_common_sheet_style(worksheet)

    headers = ["Gösterge", "Değer", "Açıklama"]
    rows = [
        [
            card["label"],
            card["value"],
            card["hint"],
        ]
        for card in report["summary_cards"]
    ]

    next_row = write_table(
        worksheet=worksheet,
        start_row=4,
        headers=headers,
        rows=rows,
        column_widths=[26, 18, 48],
    )

    highlight_headers = ["Başlık", "Sonuç"]
    highlight_rows = [
        ["En Çok Kullanılan Alan", report["most_used_area"]],
        ["En Çok Kullanılan Masa", report["most_used_table"]],
    ]

    write_table(
        worksheet=worksheet,
        start_row=next_row,
        headers=highlight_headers,
        rows=highlight_rows,
        column_widths=[28, 58],
    )

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 55


def build_area_sheet(workbook, report):
    worksheet = workbook.create_sheet("Alan Bazli Ozet")

    style_title_area(
        worksheet=worksheet,
        title="Alan Bazlı Özet",
        subtitle=f"Tarih: {report['selected_date_display']}",
        total_columns=8,
    )

    apply_common_sheet_style(worksheet)

    headers = [
        "Alan",
        "Oturum",
        "Müşteri",
        "Kapanan",
        "Toplam Süre",
        "Ortalama Süre",
        "Toplam Süre (dk)",
        "Ortalama Süre (dk)",
    ]

    rows = [
        [
            row["area_name"],
            row["session_count"],
            row["guest_count"],
            row["completed_count"],
            row["total_duration"],
            row["average_duration"],
            row["total_duration_minutes"],
            row["average_duration_minutes"] or 0,
        ]
        for row in report["area_rows"]
    ]

    write_table(
        worksheet=worksheet,
        start_row=4,
        headers=headers,
        rows=rows,
        column_widths=[24, 12, 12, 12, 18, 18, 18, 20],
    )


def build_tables_sheet(workbook, report):
    worksheet = workbook.create_sheet("En Cok Kullanilan Masalar")

    style_title_area(
        worksheet=worksheet,
        title="En Çok Kullanılan Masalar",
        subtitle=f"Tarih: {report['selected_date_display']} | İlk 10 masa",
        total_columns=9,
    )

    apply_common_sheet_style(worksheet)

    headers = [
        "Masa",
        "Alan",
        "Oturum",
        "Müşteri",
        "Kapanan",
        "Toplam Süre",
        "Ortalama Süre",
        "Toplam Süre (dk)",
        "Ortalama Süre (dk)",
    ]

    rows = [
        [
            row["table_code"],
            row["area_name"],
            row["session_count"],
            row["guest_count"],
            row["completed_count"],
            row["total_duration"],
            row["average_duration"],
            row["total_duration_minutes"],
            row["average_duration_minutes"] or 0,
        ]
        for row in report["table_rows"]
    ]

    write_table(
        worksheet=worksheet,
        start_row=4,
        headers=headers,
        rows=rows,
        column_widths=[14, 22, 12, 12, 12, 18, 18, 18, 20],
    )


def build_staff_sheet(workbook, report):
    worksheet = workbook.create_sheet("Personel Islemleri")

    style_title_area(
        worksheet=worksheet,
        title="Personel İşlem Özeti",
        subtitle=f"Tarih: {report['selected_date_display']}",
        total_columns=6,
    )

    apply_common_sheet_style(worksheet)

    headers = [
        "Personel",
        "Rol",
        "Masa Açma",
        "Masa Boşaltma",
        "Transfer",
        "Toplam İşlem",
    ]

    rows = [
        [
            row["username"],
            row["role"],
            row["assigned_count"],
            row["cleared_count"],
            row["transfer_count"],
            row["total_count"],
        ]
        for row in report["staff_rows"]
    ]

    write_table(
        worksheet=worksheet,
        start_row=4,
        headers=headers,
        rows=rows,
        column_widths=[24, 20, 14, 16, 14, 16],
    )


def build_transfer_sheet(workbook, report):
    worksheet = workbook.create_sheet("Masa Transferleri")

    style_title_area(
        worksheet=worksheet,
        title="Masa Transferleri",
        subtitle=f"Tarih: {report['selected_date_display']}",
        total_columns=8,
    )

    apply_common_sheet_style(worksheet)

    headers = [
        "Saat",
        "Kaynak Masa",
        "Kaynak Alan",
        "Hedef Masa",
        "Hedef Alan",
        "Kişi",
        "Müşteri",
        "Personel",
    ]

    rows = [
        [
            row["time"],
            row["old_table_code"],
            row["old_area_name"],
            row["new_table_code"],
            row["new_area_name"],
            row["party_size"],
            row["customer_name"],
            row["username"],
        ]
        for row in report["transfer_rows"]
    ]

    write_table(
        worksheet=worksheet,
        start_row=4,
        headers=headers,
        rows=rows,
        column_widths=[12, 16, 22, 16, 22, 10, 26, 22],
    )


def build_sessions_sheet(workbook, report):
    worksheet = workbook.create_sheet("Gunluk Masa Hareketleri")

    style_title_area(
        worksheet=worksheet,
        title="Günlük Masa Hareketleri",
        subtitle=f"Tarih: {report['selected_date_display']}",
        total_columns=10,
    )

    apply_common_sheet_style(worksheet)

    headers = [
        "Masa",
        "Alan",
        "Kişi",
        "Müşteri",
        "Telefon",
        "Giriş",
        "Çıkış",
        "Süre",
        "Süre (dk)",
        "Durum",
    ]

    rows = [
        [
            row["table_code"],
            row["area_name"],
            row["party_size"],
            row["customer_name"],
            row["customer_phone"],
            row["check_in"],
            row["check_out"],
            row["duration"],
            row["duration_minutes"] or 0,
            row["status_label"],
        ]
        for row in report["session_rows"]
    ]

    write_table(
        worksheet=worksheet,
        start_row=4,
        headers=headers,
        rows=rows,
        column_widths=[12, 22, 10, 26, 18, 12, 12, 16, 14, 16],
    )


def build_daily_summary_excel_file(report):
    workbook = Workbook()

    workbook.properties.title = "Lido Genel Gün Özeti"
    workbook.properties.subject = "Lido Masa Takip Sistemi Günlük Rapor"
    workbook.properties.creator = "Lido Masa Takip Sistemi"

    build_summary_sheet(workbook, report)
    build_area_sheet(workbook, report)
    build_tables_sheet(workbook, report)
    build_staff_sheet(workbook, report)
    build_transfer_sheet(workbook, report)
    build_sessions_sheet(workbook, report)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def find_first_existing_path(paths):
    for path in paths:
        if path and Path(path).exists():
            return str(path)

    return None


def register_pdf_fonts():
    app_root = Path(current_app.root_path)

    regular_font_path = find_first_existing_path(
        [
            app_root / "static" / "fonts" / "DejaVuSans.ttf",
            app_root / "static" / "fonts" / "Arial.ttf",
            Path("C:/Windows/Fonts/arial.ttf"),
            Path("C:/Windows/Fonts/calibri.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
        ]
    )

    bold_font_path = find_first_existing_path(
        [
            app_root / "static" / "fonts" / "DejaVuSans-Bold.ttf",
            app_root / "static" / "fonts" / "Arial Bold.ttf",
            Path("C:/Windows/Fonts/arialbd.ttf"),
            Path("C:/Windows/Fonts/calibrib.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            Path("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
        ]
    )

    if regular_font_path and bold_font_path:
        pdfmetrics.registerFont(TTFont("LidoFont", regular_font_path))
        pdfmetrics.registerFont(TTFont("LidoFont-Bold", bold_font_path))
        return "LidoFont", "LidoFont-Bold", True

    return "Helvetica", "Helvetica-Bold", False


def pdf_safe_text(value, unicode_enabled):
    if value is None:
        text = "-"
    else:
        text = str(value)

    if not unicode_enabled:
        text = text.translate(TURKISH_ASCII_TRANSLATION)

    return escape(text)


def create_pdf_styles(base_font, bold_font):
    return {
        "title": ParagraphStyle(
            "LidoTitle",
            fontName=bold_font,
            fontSize=22,
            leading=26,
            textColor=pdf_hex(COLOR_WHITE),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "subtitle": ParagraphStyle(
            "LidoSubtitle",
            fontName=base_font,
            fontSize=9.2,
            leading=12.2,
            textColor=pdf_hex(COLOR_WHITE),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "section": ParagraphStyle(
            "LidoSection",
            fontName=bold_font,
            fontSize=12.2,
            leading=15.5,
            textColor=pdf_hex(COLOR_NAVY),
            spaceBefore=2,
            spaceAfter=6,
            keepWithNext=1,
            wordWrap="CJK",
        ),
        "body": ParagraphStyle(
            "LidoBody",
            fontName=base_font,
            fontSize=7.7,
            leading=9.7,
            textColor=pdf_hex(COLOR_TEXT),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "body_center": ParagraphStyle(
            "LidoBodyCenter",
            fontName=base_font,
            fontSize=7.7,
            leading=9.7,
            textColor=pdf_hex(COLOR_TEXT),
            alignment=TA_CENTER,
            wordWrap="CJK",
        ),
        "body_right": ParagraphStyle(
            "LidoBodyRight",
            fontName=base_font,
            fontSize=7.7,
            leading=9.7,
            textColor=pdf_hex(COLOR_TEXT),
            alignment=TA_RIGHT,
            wordWrap="CJK",
        ),
        "body_bold": ParagraphStyle(
            "LidoBodyBold",
            fontName=bold_font,
            fontSize=7.9,
            leading=10,
            textColor=pdf_hex(COLOR_TEXT),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "body_bold_center": ParagraphStyle(
            "LidoBodyBoldCenter",
            fontName=bold_font,
            fontSize=7.9,
            leading=10,
            textColor=pdf_hex(COLOR_TEXT),
            alignment=TA_CENTER,
            wordWrap="CJK",
        ),
        "header_cell": ParagraphStyle(
            "LidoHeaderCell",
            fontName=bold_font,
            fontSize=7.4,
            leading=9.2,
            textColor=pdf_hex(COLOR_NAVY),
            alignment=TA_CENTER,
            wordWrap="CJK",
        ),
        "small": ParagraphStyle(
            "LidoSmall",
            fontName=base_font,
            fontSize=6.8,
            leading=8.4,
            textColor=pdf_hex(COLOR_MUTED),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "small_center": ParagraphStyle(
            "LidoSmallCenter",
            fontName=base_font,
            fontSize=6.8,
            leading=8.4,
            textColor=pdf_hex(COLOR_MUTED),
            alignment=TA_CENTER,
            wordWrap="CJK",
        ),
        "card_label": ParagraphStyle(
            "LidoCardLabel",
            fontName=bold_font,
            fontSize=7.5,
            leading=9,
            textColor=pdf_hex(COLOR_MUTED),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "card_value": ParagraphStyle(
            "LidoCardValue",
            fontName=bold_font,
            fontSize=17,
            leading=20,
            textColor=pdf_hex(COLOR_NAVY),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "card_hint": ParagraphStyle(
            "LidoCardHint",
            fontName=base_font,
            fontSize=6.6,
            leading=8.2,
            textColor=pdf_hex(COLOR_MUTED),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "highlight_label": ParagraphStyle(
            "LidoHighlightLabel",
            fontName=bold_font,
            fontSize=8,
            leading=10,
            textColor=pdf_hex(COLOR_WHITE),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "highlight_value": ParagraphStyle(
            "LidoHighlightValue",
            fontName=bold_font,
            fontSize=8.4,
            leading=10.5,
            textColor=pdf_hex(COLOR_NAVY),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "info_label": ParagraphStyle(
            "LidoInfoLabel",
            fontName=bold_font,
            fontSize=8,
            leading=10,
            textColor=pdf_hex(COLOR_NAVY),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
        "info": ParagraphStyle(
            "LidoInfo",
            fontName=base_font,
            fontSize=8,
            leading=11,
            textColor=pdf_hex(COLOR_MUTED),
            alignment=TA_LEFT,
            wordWrap="CJK",
        ),
    }


def paragraph(value, style, unicode_enabled):
    return Paragraph(pdf_safe_text(value, unicode_enabled), style)


def add_pdf_footer(canvas, doc, base_font, unicode_enabled):
    canvas.saveState()
    canvas.setFont(base_font, 7)
    canvas.setFillColor(pdf_hex(COLOR_MUTED))

    footer_text = "Lido Masa Takip Sistemi"
    if not unicode_enabled:
        footer_text = footer_text.translate(TURKISH_ASCII_TRANSLATION)

    canvas.drawString(1.2 * cm, 0.7 * cm, footer_text)
    canvas.drawRightString(
        landscape(A4)[0] - 1.2 * cm,
        0.7 * cm,
        f"Sayfa {doc.page}",
    )
    canvas.restoreState()


def build_pdf_header(report, styles, unicode_enabled):
    logo_path = Path(current_app.root_path) / "static" / "img" / "lido.png"

    title_block = [
        paragraph("Lido Genel Gün Özeti", styles["title"], unicode_enabled),
        Spacer(1, 0.09 * cm),
        paragraph(
            f"Tarih: {report['selected_date_display']} | Oluşturma: {report['generated_at']}",
            styles["subtitle"],
            unicode_enabled,
        ),
        Spacer(1, 0.03 * cm),
        paragraph(
            "Masa doluluk, müşteri hareketi ve personel işlem raporu",
            styles["subtitle"],
            unicode_enabled,
        ),
    ]

    if logo_path.exists():
        logo = Image(str(logo_path), width=2.95 * cm, height=1.29 * cm)
        logo.hAlign = "CENTER"

        logo_box = PdfTable(
            [[logo]],
            colWidths=[3.55 * cm],
            rowHeights=[1.58 * cm],
            hAlign="CENTER",
        )
        logo_box.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), pdf_hex(COLOR_WHITE)),
                    ("BOX", (0, 0), (-1, -1), 0.55, pdf_hex(COLOR_BORDER)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )

        header_data = [[logo_box, title_block]]
        column_widths = [4.1 * cm, PDF_CONTENT_WIDTH - (4.1 * cm)]
    else:
        header_data = [[title_block]]
        column_widths = [PDF_CONTENT_WIDTH]

    table = PdfTable(header_data, colWidths=column_widths, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), pdf_hex(COLOR_NAVY)),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 11),
                ("RIGHTPADDING", (0, 0), (-1, -1), 11),
                ("TOPPADDING", (0, 0), (-1, -1), 11),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 11),
                ("BOX", (0, 0), (-1, -1), 0.4, pdf_hex(COLOR_NAVY)),
                ("LINEBELOW", (0, 0), (-1, -1), 2.2, pdf_hex(COLOR_LIGHT_BLUE)),
            ]
        )
    )

    return table


def build_pdf_summary_cards(report, styles, unicode_enabled):
    card_cells = []
    column_count = 3

    for card in report["summary_cards"]:
        cell_content = [
            paragraph(card["label"], styles["card_label"], unicode_enabled),
            Spacer(1, 0.07 * cm),
            paragraph(card["value"], styles["card_value"], unicode_enabled),
            Spacer(1, 0.05 * cm),
            paragraph(card["hint"], styles["card_hint"], unicode_enabled),
        ]
        card_cells.append(cell_content)

    rows = []

    for start_index in range(0, len(card_cells), column_count):
        row = card_cells[start_index:start_index + column_count]

        while len(row) < column_count:
            row.append("")

        rows.append(row)

    table = PdfTable(
        rows,
        colWidths=[PDF_CONTENT_WIDTH / column_count] * column_count,
        hAlign="LEFT",
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), pdf_hex(COLOR_WHITE)),
                ("BOX", (0, 0), (-1, -1), 0.55, pdf_hex(COLOR_BORDER)),
                ("INNERGRID", (0, 0), (-1, -1), 0.45, pdf_hex(COLOR_BORDER)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), pdf_hex("FBFCFF")),
            ]
        )
    )

    return [
        paragraph("Yönetici Özeti", styles["section"], unicode_enabled),
        table,
        Spacer(1, 0.35 * cm),
    ]


def build_pdf_highlights(report, styles, unicode_enabled):
    data = [
        [
            paragraph("En Çok Kullanılan Alan", styles["highlight_label"], unicode_enabled),
            paragraph(report["most_used_area"], styles["highlight_value"], unicode_enabled),
        ],
        [
            paragraph("En Çok Kullanılan Masa", styles["highlight_label"], unicode_enabled),
            paragraph(report["most_used_table"], styles["highlight_value"], unicode_enabled),
        ],
    ]

    table = PdfTable(
        data,
        colWidths=[6.5 * cm, PDF_CONTENT_WIDTH - (6.5 * cm)],
        hAlign="LEFT",
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), pdf_hex(COLOR_NAVY)),
                ("BACKGROUND", (1, 0), (1, -1), pdf_hex(COLOR_LIGHT_BLUE)),
                ("GRID", (0, 0), (-1, -1), 0.45, pdf_hex(COLOR_BORDER)),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    return [
        paragraph("Öne Çıkanlar", styles["section"], unicode_enabled),
        table,
        Spacer(1, 0.35 * cm),
    ]


def build_pdf_info_box(message, styles, unicode_enabled):
    table = PdfTable(
        [
            [
                paragraph("Bilgi", styles["info_label"], unicode_enabled),
                paragraph(message, styles["info"], unicode_enabled),
            ]
        ],
        colWidths=[2.2 * cm, PDF_CONTENT_WIDTH - (2.2 * cm)],
        hAlign="LEFT",
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, 0), pdf_hex(COLOR_LIGHT_BLUE)),
                ("BACKGROUND", (1, 0), (1, 0), pdf_hex("FBFCFF")),
                ("BOX", (0, 0), (-1, -1), 0.5, pdf_hex(COLOR_BORDER)),
                ("LINEBEFORE", (0, 0), (0, 0), 2.0, pdf_hex(COLOR_NAVY)),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    return table


def build_pdf_table(title, headers, rows, column_widths, alignments, styles, unicode_enabled):
    title_flowable = paragraph(title, styles["section"], unicode_enabled)

    if not rows:
        return [
            KeepTogether(
                [
                    title_flowable,
                    build_pdf_info_box(
                        f"{title} için seçilen tarihte kayıt bulunmuyor.",
                        styles,
                        unicode_enabled,
                    ),
                    Spacer(1, 0.32 * cm),
                ]
            )
        ]

    table_data = [
        [
            paragraph(header, styles["header_cell"], unicode_enabled)
            for header in headers
        ]
    ]

    for row in rows:
        prepared_row = []

        for index, value in enumerate(row):
            alignment = alignments[index] if index < len(alignments) else "left"

            if alignment == "center":
                style = styles["body_center"]
            elif alignment == "right":
                style = styles["body_right"]
            else:
                style = styles["body"]

            prepared_row.append(paragraph(value, style, unicode_enabled))

        table_data.append(prepared_row)

    table = PdfTable(
        table_data,
        colWidths=column_widths,
        repeatRows=1,
        hAlign="LEFT",
        splitByRow=1,
    )

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), pdf_hex(COLOR_LIGHT_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), pdf_hex(COLOR_NAVY)),
        ("GRID", (0, 0), (-1, -1), 0.35, pdf_hex(COLOR_BORDER)),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, pdf_hex(COLOR_NAVY_SOFT)),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4.8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4.8),
        ("TOPPADDING", (0, 0), (-1, -1), 4.2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4.2),
    ]

    for row_index in range(1, len(table_data)):
        if row_index % 2 == 0:
            style_commands.append(
                ("BACKGROUND", (0, row_index), (-1, row_index), pdf_hex("FAFBFD"))
            )

    table.setStyle(TableStyle(style_commands))

    elements = [
        title_flowable,
        table,
        Spacer(1, 0.34 * cm),
    ]

    if len(rows) <= 9:
        return [KeepTogether(elements)]

    return elements


def build_daily_summary_pdf_file(report):
    base_font, bold_font, unicode_enabled = register_pdf_fonts()
    styles = create_pdf_styles(base_font, bold_font)

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1.05 * cm,
        leftMargin=1.05 * cm,
        topMargin=0.9 * cm,
        bottomMargin=1.0 * cm,
        title="Lido Genel Gün Özeti",
        author="Lido Masa Takip Sistemi",
    )

    elements = [
        build_pdf_header(report, styles, unicode_enabled),
        Spacer(1, 0.4 * cm),
    ]

    elements.append(
        KeepTogether(
            build_pdf_summary_cards(report, styles, unicode_enabled)
        )
    )

    elements.append(
        KeepTogether(
            build_pdf_highlights(report, styles, unicode_enabled)
        )
    )

    elements.extend(
        build_pdf_table(
            title="Alan Bazlı Özet",
            headers=["Alan", "Oturum", "Müşteri", "Kapanan", "Toplam Süre", "Ortalama"],
            rows=[
                [
                    row["area_name"],
                    row["session_count"],
                    row["guest_count"],
                    row["completed_count"],
                    row["total_duration"],
                    row["average_duration"],
                ]
                for row in report["area_rows"]
            ],
            column_widths=[7.75 * cm, 3.15 * cm, 3.15 * cm, 3.15 * cm, 5.05 * cm, 5.05 * cm],
            alignments=["left", "center", "center", "center", "center", "center"],
            styles=styles,
            unicode_enabled=unicode_enabled,
        )
    )

    elements.extend(
        build_pdf_table(
            title="En Çok Kullanılan Masalar",
            headers=["Masa", "Alan", "Oturum", "Müşteri", "Kapanan", "Toplam Süre", "Ort. Süre"],
            rows=[
                [
                    row["table_code"],
                    row["area_name"],
                    row["session_count"],
                    row["guest_count"],
                    row["completed_count"],
                    row["total_duration"],
                    row["average_duration"],
                ]
                for row in report["table_rows"]
            ],
            column_widths=[2.7 * cm, 6.4 * cm, 3.0 * cm, 3.0 * cm, 3.0 * cm, 4.6 * cm, 4.6 * cm],
            alignments=["center", "left", "center", "center", "center", "center", "center"],
            styles=styles,
            unicode_enabled=unicode_enabled,
        )
    )

    elements.extend(
        build_pdf_table(
            title="Personel İşlem Özeti",
            headers=["Personel", "Rol", "Masa Açma", "Boşaltma", "Transfer", "Toplam"],
            rows=[
                [
                    row["username"],
                    row["role"],
                    row["assigned_count"],
                    row["cleared_count"],
                    row["transfer_count"],
                    row["total_count"],
                ]
                for row in report["staff_rows"]
            ],
            column_widths=[6.8 * cm, 5.8 * cm, 3.5 * cm, 3.5 * cm, 3.5 * cm, 4.2 * cm],
            alignments=["left", "left", "center", "center", "center", "center"],
            styles=styles,
            unicode_enabled=unicode_enabled,
        )
    )

    elements.extend(
        build_pdf_table(
            title="Masa Transferleri",
            headers=["Saat", "Kaynak", "Kaynak Alan", "Hedef", "Hedef Alan", "Kişi", "Müşteri", "Personel"],
            rows=[
                [
                    row["time"],
                    row["old_table_code"],
                    row["old_area_name"],
                    row["new_table_code"],
                    row["new_area_name"],
                    row["party_size"],
                    row["customer_name"],
                    row["username"],
                ]
                for row in report["transfer_rows"]
            ],
            column_widths=[2.0 * cm, 2.6 * cm, 4.2 * cm, 2.6 * cm, 4.2 * cm, 1.6 * cm, 5.4 * cm, 4.7 * cm],
            alignments=["center", "center", "left", "center", "left", "center", "left", "left"],
            styles=styles,
            unicode_enabled=unicode_enabled,
        )
    )

    if report["session_rows"]:
        elements.append(PageBreak())

    elements.extend(
        build_pdf_table(
            title="Günlük Masa Hareketleri",
            headers=["Masa", "Alan", "Kişi", "Müşteri", "Telefon", "Giriş", "Çıkış", "Süre", "Durum"],
            rows=[
                [
                    row["table_code"],
                    row["area_name"],
                    row["party_size"],
                    row["customer_name"],
                    row["customer_phone"],
                    row["check_in"],
                    row["check_out"],
                    row["duration"],
                    row["status_label"],
                ]
                for row in report["session_rows"]
            ],
            column_widths=[2.0 * cm, 3.9 * cm, 1.5 * cm, 5.7 * cm, 3.7 * cm, 2.0 * cm, 2.0 * cm, 3.1 * cm, 3.4 * cm],
            alignments=["center", "left", "center", "left", "left", "center", "center", "center", "center"],
            styles=styles,
            unicode_enabled=unicode_enabled,
        )
    )

    doc.build(
        elements,
        onFirstPage=lambda canvas, document: add_pdf_footer(canvas, document, base_font, unicode_enabled),
        onLaterPages=lambda canvas, document: add_pdf_footer(canvas, document, base_font, unicode_enabled),
    )

    output.seek(0)

    return output


@reports_bp.route("/daily-summary", methods=["GET"])
@admin_required
def daily_summary():
    selected_date = parse_selected_date(request.args.get("date"))
    report = build_daily_summary_report(selected_date)

    return render_template(
        "admin/reports/daily_summary.html",
        app_name="Lido Masa Takip Sistemi",
        report=report,
    )


@reports_bp.route("/daily-summary/excel", methods=["GET"])
@admin_required
def daily_summary_excel():
    selected_date = parse_selected_date(request.args.get("date"))
    report = build_daily_summary_report(selected_date)
    excel_file = build_daily_summary_excel_file(report)

    filename = f"lido_genel_gun_ozeti_{report['selected_date_value']}.xlsx"

    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/daily-summary/pdf", methods=["GET"])
@admin_required
def daily_summary_pdf():
    selected_date = parse_selected_date(request.args.get("date"))
    report = build_daily_summary_report(selected_date)
    pdf_file = build_daily_summary_pdf_file(report)

    filename = f"lido_genel_gun_ozeti_{report['selected_date_value']}.pdf"

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )